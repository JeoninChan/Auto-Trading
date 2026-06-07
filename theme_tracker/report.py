import sys
import argparse
import time
import pandas as pd
import yfinance as yf
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from loguru import logger

sys.path.insert(0, str(Path(__file__).parents[1]))

from theme_tracker.classifier import classify
from theme_tracker.relevance import check_relevance, check_subsidiary_relevance
from theme_tracker.news_signal import assess
from theme_tracker.config import HOT_THEMES, DELAY, UNIVERSE_PATH

OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

BATCH_SIZE = 50
COLS = ["티커", "테마", "서브테마", "HOT테마연관도", "자회사연관도", "시총", "뉴스임박", "뉴스임박근거", "전체연관도"]


def fmt_usd(amount) -> str:
    if not amount:
        return "N/A"
    eok = int(amount // 100_000_000)
    man = int((amount % 100_000_000) // 10_000)
    if eok > 0:
        return f"{eok}억 {man:,}만달러" if man > 0 else f"{eok}억달러"
    return f"{man:,}만달러" if man > 0 else "N/A"


def get_market_cap(ticker: str) -> str:
    try:
        mc = yf.Ticker(ticker).fast_info["market_cap"]
        return fmt_usd(mc)
    except Exception:
        return "N/A"


def load_tickers(ticker: str | None, theme_filter: str | None) -> list[str]:
    if ticker:
        return [t.strip().upper() for t in ticker.split(",")]

    universe_path = Path(__file__).parents[1] / UNIVERSE_PATH
    if not universe_path.exists():
        logger.error(f"유니버스 파일 없음: {universe_path}")
        sys.exit(1)

    df = pd.read_parquet(universe_path)
    tickers = [t.upper() for t in df["ticker"].tolist()]
    logger.info(f"처리 대상: {len(tickers)}개 종목")
    return tickers


def append_excel(rows: list[dict], path: Path, first_batch: bool):
    if first_batch or not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "테마추적"
        ws.append(COLS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
    else:
        wb = load_workbook(path)
        ws = wb.active

    for row in rows:
        ws.append([row.get(c, "") for c in COLS])

    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col) + 4
        ws.column_dimensions[col[0].column_letter].width = min(width, 50)

    wb.save(path)


def process_ticker(ticker: str) -> dict:
    theme, sub = classify(ticker)

    rel_map = {}
    hot_rel = "연관 없음"
    for ht in HOT_THEMES:
        rel = check_relevance(ticker, ht)
        rel_map[ht] = rel
        if rel in ("직접 연관", "간접 연관") and hot_rel == "연관 없음":
            hot_rel = f"{ht} ({rel})"

    # 자회사 연관도: HOT_THEMES 중 가장 강한 결과 반환
    sub_rel = "연관 없음"
    for ht in HOT_THEMES:
        sr = check_subsidiary_relevance(ticker, ht)
        if sr == "직접 연관":
            sub_rel = f"{ht} (직접 연관)"
            break
        if sr == "간접 연관" and sub_rel == "연관 없음":
            sub_rel = f"{ht} (간접 연관)"

    news = assess(ticker)

    return {
        "티커":          ticker,
        "테마":          theme,
        "서브테마":      sub,
        "HOT테마연관도": hot_rel,
        "자회사연관도":  sub_rel,
        "시총":          get_market_cap(ticker),
        "뉴스임박":      news["뉴스임박"],
        "뉴스임박근거":  news["뉴스임박근거"],
        "전체연관도":    " | ".join(f"{k}: {v}" for k, v in rel_map.items()),
    }


def main():
    parser = argparse.ArgumentParser(description="테마주 분류 + HOT테마 연관도 + 뉴스 임박 감지")
    parser.add_argument("--ticker", "-t", default=None, help="특정 티커 (쉼표 구분: LUNR,ASTS)")
    parser.add_argument("--theme",  "-m", default=None, help="특정 테마 필터 (예: 우주)")
    args = parser.parse_args()

    tickers = load_tickers(args.ticker, args.theme)
    total = len(tickers)

    date_str = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = OUTPUT_DIR / f"theme_tracker_{date_str}.xlsx"

    batch: list[dict] = []
    highlights: list[dict] = []
    first_batch = True

    for i, ticker in enumerate(tickers, 1):
        logger.info(f"[{i}/{total}] {ticker}")
        try:
            row = process_ticker(ticker)
        except Exception as e:
            logger.debug(f"{ticker} 실패: {e}")
            row = {c: "" for c in COLS}
            row["티커"] = ticker

        batch.append(row)
        if "직접 연관" in row["HOT테마연관도"] and row["뉴스임박"] == "높음":
            highlights.append(row)

        if len(batch) >= BATCH_SIZE or i == total:
            append_excel(batch, out_path, first_batch)
            logger.info(f"중간 저장 ({i}/{total}) → {out_path.name}")
            batch.clear()
            first_batch = False

        time.sleep(DELAY)

    if highlights:
        print(f"\n{'='*60}")
        print(f"  주목 종목 (직접 연관 + 뉴스 임박 높음)")
        print(f"{'='*60}")
        for r in highlights:
            print(f"  {r['티커']:8s} | {r['테마']} > {r['서브테마']}")
            print(f"           근거: {r['뉴스임박근거']}")
        print(f"{'='*60}\n")

    logger.success(f"완료 → {out_path}")


if __name__ == "__main__":
    main()
